from openpyxl import load_workbook


class ExcelParserDeaths:
    BLOOD_CIRCULATION_ID = 'IX'

    # ids that are contained in each xlsx file
    COMMON_IDS = ['I', 'II', 'III', 'IV', 'V', 'VI', 'IX', 'X', 'XI', 'XII', 'XIII', 'XIV', 'XVI', 'XVII', 'XVIII', 'XX']

    def __init__(self, region_name):
        self.region_name = region_name

        self.wb = load_workbook(filename="data/excel/{}.xlsx".format(region_name), read_only=True)
        self.sheet = self.wb.active

    def findRowById(self, row_id):
        for row in self.sheet.rows:
            if row[0].value == row_id:
                return row

        return False

    def getValuesByYears(self, row, start_year = 2005, end_year = 2014):
        result_list = []

        for year in range(start_year, end_year + 1):
            year_value = row[year - 2003].value
            if year_value == '- ':
                result_list.append(0)
            else:    
                result_list.append(year_value)

        return result_list

    def getValuesByYearsForCommonIds(self):
        result = dict()

        for id in self.COMMON_IDS:
            row = self.findRowById(id)
            result[id] = self.getValuesByYears(row)

        return result


    def close(self):
        self.wb.close()


# # a helper method
# def getCountForYear(row, year: int):
#     assert year >= 2005 and year <= 2014
#     return row[year - 2003].value

# def getValuesByYears(region_name, row_id):
#     wb = load_workbook(filename="data/excel/{}.xlsx".format(region_name), read_only=True)
#     sheet = wb.active

#     result_list = []

#     for row in sheet.rows:
#         if row[0].value == row_id:
#             for year in range(2005,2015):
#                 result_list.append(getCountForYear(row,year))

#             # we have reached the target row, so skip the rest
#             break

#     wb.close()

#     return result_list

# def hasRowWithValue(region_name, row_id):
#     wb = load_workbook(filename="data/excel/{}.xlsx".format(region_name), read_only=True)
#     sheet = wb.active

#     for row in sheet.rows:
#         if row[0].value == row_id:
#             return True

#     return False


# def getCirulationValuesByYears(region_name):
#     return getValuesByYears(region_name, BLOOD_CIRCULATION_ID)
